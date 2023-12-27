import openpyxl
import re
from openpyxl.styles import PatternFill


# регулярное выражение для выделения групп из совпадений 
reg_groups = r"(?P<date_document>\d{2}\.\d{2}\.\d{2})[\s|\\n]*(?P<operation>[А-яЁё]{0,7})[\s|\\n]*\((?P<number_document>\d{0,5})[\s|\\n]*от[\s|\\n]\d{2}[\s|\\n]*\.\d{2}\.\d{4}\)[\s|\\n]*(?P<amount_invoice>\d{0,4}[\s|\\n]*\d{3}\,\d{2})"

def color_xlsx_cell(list_matches_pdf_file, xlsx_file):
    wb = openpyxl.reader.excel.load_workbook(filename=xlsx_file)
    wb.active = 0
    sheet = wb.active
    iter_list = []  # Для исключения уже заграшеных строк
    for match in list_matches_pdf_file:
        # Разбиваем на группы каждый элемент из list_matches_pdf_file
        # Группы =>  date_document, number_document, amount_invoice
        dict_match = re.search(reg_groups, match).groupdict()
        for row in sheet.iter_rows():  # итерируемся по каждой строке в xlsx_file
            pdf_date = dict_match['date_document']  # Даты не используются
            pdf_number = dict_match['number_document']
            pdf_amount = dict_match['amount_invoice'].replace(' ', '').split(",")[0].split(",")[0]
            xlsx_column_1 = row[1].value  # Значение колонка номера строки
            xlsx_column_3 = row[3].value  # Значение колонка Документ
            xlsx_column_4 = row[4].value  # Значение колонка Дебет
            xlsx_column_5 = row[5].value  # Зничение колонка Кредит
            if isinstance(xlsx_column_1, int) and xlsx_column_1 not in iter_list:
                if pdf_number in xlsx_column_3 and (int(pdf_amount) == xlsx_column_4 or int(pdf_amount) == xlsx_column_5):
                    iter_list.append(xlsx_column_1)
                    row[3].fill = PatternFill(fill_type='solid', start_color='00FF00')
                    row[4].fill = PatternFill(fill_type='solid', start_color='00FF00')
                    row[5].fill = PatternFill(fill_type='solid', start_color='00FF00')
                    break
                elif pdf_number in xlsx_column_3 and (int(pdf_amount) != xlsx_column_4 or int(pdf_amount) != xlsx_column_5):
                    iter_list.append(xlsx_column_1)
                    row[3].fill = PatternFill(fill_type='solid', start_color='FF9999')
                    row[4].fill = PatternFill(fill_type='solid', start_color='FF9999')
                    row[5].fill = PatternFill(fill_type='solid', start_color='FF9999')
                elif pdf_number not in xlsx_column_3:
                    print(pdf_number)
                # Выход из внутреннего цикла и переход к следующему match из list_matches_pdf_file
                break  
    # Итерируемся по ячейкам, ищем не закрашенные => это значит что данная строка есть только в xlsx_file
    for row in sheet.iter_rows():  # итерируемся по кождой строке в xlsx_file
        xlsx_column_1 = row[1].value  # Значение колонка номера строки
        xlsx_column_3 = row[3].fill  # Значение колонка Документ
        if isinstance(xlsx_column_1, int) and row[3].fill.fgColor.value == "00000000":
            row[3].fill = PatternFill(fill_type='solid', start_color='FFFF99')
            row[4].fill = PatternFill(fill_type='solid', start_color='FFFF99')
            row[5].fill = PatternFill(fill_type='solid', start_color='FFFF99')
    # Пересохраняем xlsx_file
    wb.save(xlsx_file)
    


if __name__ == '__main__':
    list_matches_pdf_file = ['27.01.23\\nОплата (120 от 27.01.2023)\\n32 000,00', '31.01.23\\nПродажа (26 от 31.01.2023)\\n32 000,00', '28.02.23\\nПродажа (61 от 28.02.2023)\\n32 000,00', '01.03.23\\nОплата (283 от 01.03.2023)\\n32 000,00', '29.03.23\\nОплата (447 от 29.03.2023)\\n32 000,00', '31.03.23\\nПродажа (98 от 31.03.2023)\\n32 000,00', '30.04.23\\nПродажа (134 от 30.04.2023)\\n32 000,00', '10.05.23\\nОплата (627 от 10.05.2023)\\n32 000,00', '31.05.23\\nПродажа (171 от 31.05.2023)\\n32 000,00', '08.06.23\\nОплата (786 от 08.06.2023)\\n32 000,00', '30.06.23\\nПродажа (204 от 30.06.2023)\\n32 000,00', '11.07.23\\nОплата (971 от 11.07.2023)\\n32 000,00', '31.07.23\\nПродажа (243 от 31.07.2023)\\n32 000,00', '17.08.23\\nОплата (1130 от 17.08.2023)\\n32 000,00', '18.08.23\\nОплата (1136 от 18.08.2023)\\n32 000,00', '31.08.23\\nПродажа (277 от 31.08.2023)\\n32 000,00']
    xlsx_file = 'aiogram-bot\chat_files\акт эф решение.xlsx'
    
    color_xlsx_cell(list_matches_pdf_file, xlsx_file)